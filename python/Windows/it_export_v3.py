# -*- coding:utf8 - * -
# 导出资产管理列表
import MySQLdb
import sys, os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import chardet
import time, datetime
import string


class Excel():
    def __init__(self, config):
        self.Head = config['Head']
        self.pages = config['Sheets']['pages']
        self.default_flied = config['Sheets']['default_flied']
        self.default_flied_style = config['Sheets']['default_flied_style']
        self.Sheets = []

        self.wb = Workbook()
        # 创建sheet 生成page
        for num in range(len(self.pages)):
            title = self.pages[num]['name']
            if num == 0:
                ws = self.wb.active
                ws.title = title
            else:
                ws = self.wb.create_sheet()
                ws.title = title
            ws.sheet_properties.tabColor = "1072BA"
            self.Sheets.append(ws)

    def writeContent(self):

        for page in self.pages:
            content = []
            keyword = ''
            flied = ''
            data = []

            name = page['name']
            keyword = page['keyword']
            # print keyword
            # 游标指向对应名称sheet
            ws = self.wb[name]
            # 获取字段 没有设置则使用初始化字段
            if page.has_key('flied'):
                flied = page['flied']
            else:
                flied = self.default_flied

            # 处理首行
            end_column = len(flied)
            # 合并首行
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
            # 设置首行样式
            head_styles = self.Head['style']
            style_types = self.Head['style'].keys()
            for style_type in style_types:
                condition = ','.join(['%s=%s' % (c[0], c[1]) for c in head_styles[style_type].items()])
                # style_cmd = 'ws.row_dimensions[1].%s = %s(%s)' % (style_type,style_type.title(),condition)
                style_cmd = 'ws["A1"].%s = %s(%s)' % (style_type, style_type.title(), condition)
                exec (style_cmd)

            # 写入首行内容
            ws["A1"] = self.Head['content']

            # 获取表头字段
            title = [nameMapper[t] for t in flied]
            # 设置表头样式
            if page.has_key('flied_style'):
                flied_style = page['flied_style']
            else:
                flied_style = self.default_flied_style

            style_types = flied_style.keys()
            for style_type in style_types:
                for column in string.uppercase[:len(flied)]:
                    condition = ','.join(['%s=%s' % (c[0], c[1]) for c in flied_style[style_type].items()])
                    # style_cmd = 'ws["B%s"].%s = %s(%s)' % (column,style_type,style_type.title(),condition)
                    style_cmd = 'ws["%s2"].%s = %s(%s)' % (column, style_type, style_type.title(), condition)
                    data_cmd = 'ws["%s2"]  =  "%s"' % (column, title[string.uppercase.index(column)])
                    exec (style_cmd)
                    exec (data_cmd)

            # 获取数据
            if 'other' == keyword:
                sql = 'select %s from it where it_serial not regexp "%s(.*)"' % (
                ','.join(flied), '|'.join([k['keyword'] for k in self.pages[:-1]]))
            elif 'new' == keyword:
                sql = "select %s from it where it_dec regexp '.*新增' and it_buytime >= '%s' ;" % (
                ','.join(flied), page['start_date'])
            elif 'history_new' == keyword:
                sql = "select %s from it where it_dec regexp '.*新增' order by it_buytime;" % ','.join(flied)
            else:
                sql = "select %s from it where it_serial regexp '%s'" % (','.join(flied), keyword)

            data = self.get_mysql(sql)
            if data:
                for num in range(len(data)):
                    content.append([data[num][f] for f in flied])
                # content.append(data[num].values())

            # 写入数据
            for row in iter(content):
                ws.append(row)

    def save(self, filename='测试', suffix='.xlsx'):
        ''' 默认生成传入文件名加导出时间和.xlsx的后缀 返回文件存储位置 '''

        export_time = time.strftime("%Y%m%d", time.localtime())
        filename = filename + export_time + suffix

        if sys.platform == 'win32':
            self.wb.save(filename.decode('utf8'))
        else:
            self.wb.save(filename)
        return os.getcwd() + os.sep + filename

    def get_mysql(self, sql, host='10.10.41.2', user='root', password='test', database='jd_oss'):
        try:
            db = MySQLdb.connect(host, user, password, database, charset='utf8')
            cur = db.cursor(cursorclass=MySQLdb.cursors.DictCursor)
            cur.execute(sql)
            mysql_result = cur.fetchall()
        except Exception, e:
            print e
            mysql_result = None
        cur.close()
        db.close()
        return mysql_result


def get_last_month_fday():
    y, m, d = datetime.date.today().year, datetime.date.today().month, datetime.date.today().day
    # 如果是1月
    if m == 1:
        start_date = datetime.date(y - 1, 12, 1)
    else:
        start_date = datetime.date(y, m - 1, 1)
    return start_date


if __name__ == '__main__':
    start_date = get_last_month_fday()
    print start_date
    sys.exit()

    config = {
        'Head': {
            'content': u'上海际动网络科技股份公司 资产盘点表',
            'style'	: {
						'font':
								{
								 'size'	:	'20',
							 	 'bold' : 	True
							 	},
						'alignment' :{
								'vertical'	 :	'"center"',
								'horizontal' :	'"center"'
								}
						},
					},
			'Sheets' : {
					'default_flied' : ['it_type' ,'it_serial' ,'it_size' ,'it_dec' ,'it_cpu' ,'it_mem' ,'it_dep' ,'it_user'
                              ,'it_reason' ,'it_detail' ,'it_usetime' ,'it_buytime' ,'it_cost' ,'it_channel'],
						'default_flied_style' : {
									'font': {
										'bold' : True,
										'size' : '12'
									} ,
								},

					'pages' :  [
						{
							'name'	:	u'电脑',
							'keyword':  'SHJDPC|SHJDNB'
						},
						{
							'name' 	:	u'显示器',
							'keyword':	'SHJDLE'
						},
						{
							'name' 	:	u'测试机',
							'keyword':	'SHJDBG|SHJDSJ|SHIDBG'
						} ,
						{
                            'name'	:	u'机房',
							'keyword':	'SHJDJF'
						},
						{
							'name'	:	u'其他',
							'keyword':	'other'
						},
						{
							'name'	:	u'新增',
							'keyword':	'new',
							'flied' :   ['it_dec' ,'it_type' ,'it_serial' ,'it_size' ,'it_detail' ,'it_cost'],
							'start_date' : start_date
						},
						{
							'name'	:	u'历史所有新增',
							'keyword':	'history_new',
							'flied' :   ['it_dec' ,'it_type' ,'it_serial' ,'it_size' ,'it_detail' ,'it_cost'],
						}
					]

			}
	}

	nameMapper = {
		'it_id'			: '编号',
		'it_serial'		: '新资产编号',
		'it_type'		: '资产种类',
		'it_size'		: '型号/尺寸',
		'it_brand'		: '品牌',
		'it_os'			: '系统',
		'it_cpu'		: 'CPU(处理器)',
		'it_mem'		: '内存',
		'it_ip'			: 'IP',
		'it_mac'		: 'MAC地址',
		'it_dep'		: '部门',
		'it_user'		: '使用者',
		'it_status'		: '状态',
		'it_reason'		: '使用原因',
		'it_usetime'	: '使用时间',
		'it_buytime'	: '购买时间',
		'it_cost'		: '价格',
		'it_channel'	: '购买渠道',
		'it_detail'		: '备注',
		'it_dec'		: '详细信息'
	}

   	''' MySQL默认连接参数：
			    host = '10.10.41.2'
			    user = 'root'
			    password = 'test'
			    database = 'jd_oss'
   	'''
	e = Excel(config)
	e.writeContent()
	# 可以手动输入文件名 save(filename=filename)
	# 默认文件名 : 测试
  	print '文件保存位置:' + e.save(filename='资产盘点表')
