#-*- coding:utf8 -*-
from datetime import date,datetime,timedelta
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,Font,Color,PatternFill,Border,Side,style,Style
import time,calendar
import sys
import string

def get_month_range(month=date.today().month):
	monthRange = []
	# weekAndDay = ()

	start_date = date.today().replace(month=month,day=1)
	days_in_month = calendar.monthrange(start_date.year,start_date.month)
	for day in range(1,days_in_month[1] + 1):
		week = start_date.replace(day=day).isoweekday()
		day = start_date.replace(day=day).day
		monthRange.append((week,day))
	return month,monthRange

month,monthRange = get_month_range()


def toExcel(month=date.today().month,filename=u'工作计划.xlsx'):
	weekMapper  =  {
			1 : '一',
			2 : '二',
			3 : '三',
			4 : '四',
			5 : '五',
			6 : '六',
			7 : '日'
	}

	month,monthRange = get_month_range(month)
	# sheet_title = str(month) +  u'月份'
	sheet_title = str(month) +  u'月份'
	wb = load_workbook(filename=filename)
	wb.create_sheet(title=sheet_title)

	#首行
	ws = wb.get_sheet_by_name(sheet_title)
	# 合并
	for i in range(1,5):
		ws.merge_cells(start_row=1,start_column=i,end_row=2,end_column=i)

	ws.merge_cells(start_row=3,start_column=1,end_row=16,end_column=1)
	ws.merge_cells(start_row=17,start_column=1,end_row=28,end_column=1)
	#固定字段
	ws['A1'] = u'工作内容'
	ws['A3'] = 'IT'
	ws['A17'] = u'运维'
	ws['B1'] = u'细项'
	ws['C1'] = u'责任人'
	ws['D1'] = u'重要性'

	#获取日期位置
	column_list = [ c for c in string.uppercase ]
	#生成首行column字段
	column_list.extend([ i+j for i in column_list for j in column_list])
	# column 元素位置往右偏移一位
	# column_list.insert(0,1)

	for num in range(len(monthRange)):
		week = monthRange[num][0]
		day = monthRange[num][1]
		if week > 5 :
			print '%s.%s is 星期%s' % (month,day,weekMapper[week])
			

		#日期 
		day_row_pos = column_list[4+num] + '1'
		week_row_pos = column_list[4+num] + '2'
		ws[day_row_pos] = day
		ws[week_row_pos] = weekMapper[week]

		#设定样式
		for cell in ws.rows[:2]:
			for c in cell:
				c.fill=PatternFill(patternType='solid', fgColor=Color(rgb='FFFF00'))
				c.alignment = Alignment(vertical='center',horizontal='center') 
				c.font = Font(bold=True,size=11)
				c.border = Border(left=Side(border_style='thin',
				                           color='000000'),
				                 right=Side(border_style='thin',
				                            color='000000'),
				                 top=Side(border_style='thin',
				                          color='000000'),
				                 bottom=Side(border_style='thin',
				                             color='000000'),
				                 diagonal=Side(border_style='thin',
				                               color='000000'),
				                 diagonal_direction=0,
				                 outline=Side(border_style='thin',
				                              color='000000'),
				                 vertical=Side(border_style='thin',
				                               color='000000'),
				                 horizontal=Side(border_style='thin',
				                                color='000000')
				                )
	# map(lambda s: s.alignment=Alignment(vertical='center',horizontal='center'), 
	for cell in [ws['A3'],ws['A17']]:
		cell.alignment=Alignment(vertical='center',horizontal='center')
		cell.font = Font(bold=False,size=12)
 	wb.save('1.xlsx')

#默认为当前月份  可以手动指定月份 toExcel(month=1)
toExcel(month=11)

